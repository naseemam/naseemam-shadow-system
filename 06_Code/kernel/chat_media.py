"""Persistent business-chat attachments for Ameer Shadow System.

The store keeps user-provided files outside the repository checkout, exposes a
small safe manifest to the chat, and extracts useful context from price and
inventory files.  It does not create a governance approval: uploads are inputs
to an existing Shadow asset and remain auditable by their content hash.
"""
from __future__ import annotations

import csv
import hashlib
import json
import mimetypes
import os
import re
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable


MAX_UPLOAD_BYTES = 50 * 1024 * 1024
TEXT_PREVIEW_BYTES = 24_000
_ALLOWED_SUFFIXES = {
    ".csv", ".tsv", ".xlsx", ".xls", ".json", ".txt", ".md", ".pdf", ".doc", ".docx",
    ".png", ".jpg", ".jpeg", ".webp", ".gif", ".svg",
    ".mp4", ".mov", ".webm", ".mkv",
    ".mp3", ".wav", ".m4a", ".ogg", ".aac", ".webm",
}


def _now() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def _safe_filename(value: str) -> str:
    name = Path(str(value or "attachment")).name.strip()
    name = re.sub(r"[^\w.\-\u0600-\u06FF]+", "_", name, flags=re.UNICODE).strip("._")
    return name[:140] or "attachment"


def _category(content_type: str, suffix: str) -> str:
    mime = str(content_type or "").lower()
    suffix = suffix.lower()
    if mime.startswith("image/") or suffix in {".png", ".jpg", ".jpeg", ".webp", ".gif", ".svg"}:
        return "image"
    if mime.startswith("video/") or suffix in {".mp4", ".mov", ".webm", ".mkv"}:
        return "video"
    if mime.startswith("audio/") or suffix in {".mp3", ".wav", ".m4a", ".ogg", ".aac"}:
        return "audio"
    if suffix in {".csv", ".tsv", ".xlsx", ".xls"}:
        return "spreadsheet"
    if suffix in {".txt", ".md", ".json"}:
        return "text"
    return "document"


class ChatMediaStore:
    """Store and retrieve business-chat attachments under the persistent data root."""

    def __init__(self, data_root: str | Path) -> None:
        self.root = Path(data_root).resolve() / ".ameer" / "chat_uploads"
        self.root.mkdir(parents=True, exist_ok=True)

    def _metadata_path(self, attachment_id: str) -> Path:
        return self.root / f"{attachment_id}.json"

    def _assert_identifier(self, attachment_id: str) -> str:
        value = str(attachment_id or "").strip().lower()
        if not re.fullmatch(r"[a-f0-9]{24}", value):
            raise ValueError("invalid_attachment_id")
        return value

    def save(self, *, filename: str, content_type: str, data: bytes) -> dict[str, Any]:
        if not data:
            raise ValueError("empty_attachment")
        if len(data) > MAX_UPLOAD_BYTES:
            raise ValueError("attachment_exceeds_50mb_limit")

        display_name = _safe_filename(filename)
        suffix = Path(display_name).suffix.lower()
        if suffix not in _ALLOWED_SUFFIXES:
            raise ValueError("unsupported_attachment_type")

        attachment_id = uuid.uuid4().hex[:24]
        stored_name = f"{attachment_id}{suffix}"
        target = self.root / stored_name
        target.write_bytes(data)
        mime = str(content_type or mimetypes.guess_type(display_name)[0] or "application/octet-stream")
        metadata = {
            "attachment_id": attachment_id,
            "filename": display_name,
            "stored_name": stored_name,
            "mime_type": mime,
            "category": _category(mime, suffix),
            "size_bytes": len(data),
            "sha256": hashlib.sha256(data).hexdigest(),
            "uploaded_at": _now(),
        }
        self._metadata_path(attachment_id).write_text(json.dumps(metadata, ensure_ascii=False, indent=2), encoding="utf-8")
        return metadata

    def get(self, attachment_id: str) -> dict[str, Any]:
        attachment_id = self._assert_identifier(attachment_id)
        path = self._metadata_path(attachment_id)
        if not path.exists():
            raise FileNotFoundError("attachment_not_found")
        metadata = json.loads(path.read_text(encoding="utf-8"))
        stored = self.root / str(metadata.get("stored_name") or "")
        if not stored.is_file() or not stored.resolve().is_relative_to(self.root):
            raise FileNotFoundError("attachment_payload_not_found")
        metadata["path"] = str(stored)
        return metadata

    def public(self, metadata: dict[str, Any]) -> dict[str, Any]:
        return {
            "attachment_id": metadata["attachment_id"],
            "filename": metadata["filename"],
            "mime_type": metadata["mime_type"],
            "category": metadata["category"],
            "size_bytes": metadata["size_bytes"],
            "uploaded_at": metadata["uploaded_at"],
            "download_url": f"/chat/uploads/{metadata['attachment_id']}",
        }

    def payload_path(self, attachment_id: str) -> Path:
        return Path(self.get(attachment_id)["path"])

    def attachment_context(self, attachment_ids: Iterable[str]) -> tuple[str, list[dict[str, Any]]]:
        seen: set[str] = set()
        entries: list[dict[str, Any]] = []
        parts: list[str] = []
        remaining = TEXT_PREVIEW_BYTES
        for raw_id in attachment_ids:
            if not isinstance(raw_id, str) or raw_id in seen:
                continue
            seen.add(raw_id)
            metadata = self.get(raw_id)
            entries.append(self.public(metadata))
            label = f"- {metadata['filename']} ({metadata['category']}, {metadata['mime_type']}, {metadata['size_bytes']} bytes)"
            preview = self._preview_text(metadata, limit=remaining)
            if preview:
                clipped = preview[:remaining]
                remaining -= len(clipped)
                label += f"\n  المحتوى المستخرج:\n{clipped}"
            parts.append(label)
        if not parts:
            return "", []
        return "\n\n[مرفقات أضافها المالك إلى هذه المهمة]\n" + "\n".join(parts) + "\n[/مرفقات]", entries

    def _preview_text(self, metadata: dict[str, Any], *, limit: int) -> str:
        if limit <= 0:
            return ""
        path = Path(metadata["path"])
        suffix = path.suffix.lower()
        try:
            if suffix in {".txt", ".md", ".json", ".csv", ".tsv"}:
                return path.read_text(encoding="utf-8", errors="replace")[:limit]
            if suffix in {".xlsx", ".xls"}:
                try:
                    from openpyxl import load_workbook  # type: ignore
                except Exception:
                    return "ملف جداول مرفق. تتوفر المعالجة الكاملة بعد تهيئة مكتبة الجداول في البيئة."
                workbook = load_workbook(path, read_only=True, data_only=True)
                sheet = workbook.active
                rows = []
                for row_index, row in enumerate(sheet.iter_rows(values_only=True)):
                    if row_index >= 30:
                        break
                    values = ["" if value is None else str(value) for value in row[:16]]
                    rows.append("\t".join(values).rstrip())
                workbook.close()
                return "\n".join(rows)[:limit]
        except Exception:
            return "تعذر استخراج نص من هذا المرفق، لكنه محفوظ ومتاح لأمير ضمن المهمة."
        return ""

    def audit_summary(self, attachment_ids: Iterable[str]) -> list[dict[str, Any]]:
        summary = []
        for attachment_id in attachment_ids:
            try:
                metadata = self.get(attachment_id)
            except (ValueError, FileNotFoundError):
                continue
            summary.append({
                "attachment_id": metadata["attachment_id"],
                "filename": metadata["filename"],
                "category": metadata["category"],
                "sha256": metadata["sha256"],
            })
        return summary
